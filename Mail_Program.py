#NOTE:--->>> before you used this program you have to make sure that Email_ID, Tags and Status column with content(Email_ID,Tags)
#should be present in excel file


# Python program to read an excel file

# import openpyxl module
import openpyxl
import yagmail
# Give the location of the file
path = "Path for excel file"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active

# Cell objects also have row, column,
# and coordinate attributes that provide
# location information for the cell.

# Note: The first row or
# column integer is 1, not 0.

# Cell object is created by using
# sheet object's cell() method
bulk_dict = {}
print("Before using this programe firstly open the browser and login the gmail ID which will used in this programe \n then click this link and allow the less secure apps \n https://myaccount.google.com/lesssecureapps\n\n")
Gmail_ID = input('This Program  only used By Gmail ID So Please Enter your Gmail ID:-->> \n').lower().strip()
Password = input('Please Enter your password:-->> \n')
for var in range(sheet_obj.max_row-1):
    cell_obj = sheet_obj.cell(row = var+2, column = 1)
    cell_obj1 = sheet_obj.cell(row = var+2, column = 2)
    # Print value of cell object
    # using the value attribute
    #print(cell_obj.value+'\t'+cell_obj1.value)
    Email_ID = cell_obj.value
    #print(Email_ID)
    Tags = cell_obj1.value
    if Tags not in bulk_dict:
        temp = list()
        temp.append(Email_ID)
        bulk_dict[Tags] = temp
    else:
        bulk_dict[Tags].append(Email_ID)
list_of_keys = list(bulk_dict.keys())
subject_list = []
body_list = []
for var in range(len(list_of_keys)):
    print('For '+list_of_keys[var]+'Tag --->>\n')
    subject = input("Enter the Subject Here: ")
    message = input("Enter the message Here: \n\n")
    subject_list.append(subject)
    body_list.append(message)

final_dict = {}
s1 = len(list_of_keys)
s = 0
row_value = 2
for key in list_of_keys:
    if s1 > s:
        for var in range(len(bulk_dict[key])):
            try:
                #initializing the server connection
                yag = yagmail.SMTP(user=Gmail_ID, password=Password)
                #sending the email
                yag.send(to=bulk_dict[key][var], subject=subject_list[s], contents=body_list[s])
                temp_var = sheet_obj[f'C{row_value}']
                temp_var.value = "Email sent successfully"
                print("Email sent successfully")
                row_value+=1
            except:
                temp_var = sheet_obj[f'C{row_value}']
                temp_var.value = "Error, Email was not sent"
                row_value+=1
                print("Error, Email was not sent")
        s +=1
# the save() workbook method.
wb_obj.save("Path for excel file")

