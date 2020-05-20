# Python program to read an excel file
from requests import *
from bs4 import *
# import openpyxl module
import openpyxl

# Give the location of the file
path = "C:\\Users\\resurrrector\\Downloads\\slt.xlsx"

# To open the workbook
# workbook object is created
wb_obj = openpyxl.load_workbook(path)

# Get workbook active sheet object
# from the active attribute
sheet_obj = wb_obj.active
row_no = sheet_obj.max_row

# Ste - 1:---> count

count = 1
for j in range(2, row_no + 1):
    cell_obj = sheet_obj.cell(row=j, column=2)
    if string == 'Error':
        count += 1
    if str(string).isdecimal():
        count += 1
    else:
        break

for i in range(count + 1, row_no + 1):
    # For row 0 and column 0
    cell_obj = sheet_obj.cell(row=i, column=1)
    facebooke_page_url = cell_obj.value

    website_url = 'https://findmyfbid.com/'
    data = {'url': facebooke_page_url}
    response = post(website_url, json=data)
    # print(response.text)
    if len(response.text) == 2:
        result = 'Error'
    else:
        response1 = response.text[6:]
        result = str(response1[:-1])

    c2 = sheet_obj.cell(row=i, column=2)
    c2.value = result

    wb_obj.save("C:\\Users\\resurrrector\\Downloads\\slt.xlsx")